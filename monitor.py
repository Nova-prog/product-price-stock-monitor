from __future__ import annotations

import argparse
import csv
import datetime as dt
import html
import json
import os
import re
import sys
import tempfile
import time
import urllib.error
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


APP_NAME = "product-price-stock-monitor"
DEFAULT_PRODUCTS = "products.csv"
DEFAULT_OUTPUT_DIR = "data"
DEFAULT_WORKBOOK = "product_monitor.xlsx"
USER_AGENT = (
    "Mozilla/5.0 (compatible; product-price-stock-monitor/1.0; "
    "+https://github.com/Nova-prog/product-price-stock-monitor)"
)


PRODUCT_COLUMNS = [
    "enabled",
    "group",
    "category",
    "site",
    "name",
    "url",
    "product_code",
    "manufacturer",
    "manufacturer_code",
    "jan",
    "cas",
    "grade",
    "capacity",
    "unit",
    "quantity",
    "quantity_unit",
    "expected_price_type",
    "expected_tax_status",
    "notes",
]

LATEST_COLUMNS = [
    "checked_at",
    "result_state",
    "field_status",
    "group",
    "category",
    "site",
    "name",
    "product_code",
    "manufacturer",
    "manufacturer_code",
    "jan",
    "cas",
    "grade",
    "capacity",
    "unit",
    "quantity",
    "quantity_unit",
    "price",
    "currency",
    "price_type",
    "tax_status",
    "unit_price",
    "unit_price_basis",
    "stock_status",
    "delivery_status",
    "dispatch_status",
    "raw_price_text",
    "raw_stock_text",
    "raw_delivery_text",
    "title",
    "url",
    "error",
    "source_checked_at",
]

HISTORY_COLUMNS = LATEST_COLUMNS + ["previous_success_price", "price_diff", "price_diff_percent"]

COMPARE_COLUMNS = [
    "checked_at",
    "group",
    "category",
    "best_site",
    "best_name",
    "best_price",
    "best_unit_price",
    "currency",
    "price_type",
    "tax_status",
    "comparison_note",
    "candidate_count",
]

ERROR_COLUMNS = [
    "checked_at",
    "site",
    "name",
    "url",
    "failed_field",
    "reason",
]


SITE_ALIASES = {
    "monotaro": "monotaro",
    "モノタロウ": "monotaro",
    "askul": "askul",
    "アスクル": "askul",
    "kanto": "kanto",
    "kanto_chemical": "kanto",
    "関東化学": "kanto",
    "cica": "kanto",
    "wako": "wako",
    "fuji_wako": "wako",
    "富士フイルム和光": "wako",
    "和光": "wako",
}


@dataclass
class Product:
    enabled: bool
    group: str
    category: str
    site: str
    name: str
    url: str
    product_code: str = ""
    manufacturer: str = ""
    manufacturer_code: str = ""
    jan: str = ""
    cas: str = ""
    grade: str = ""
    capacity: str = ""
    unit: str = ""
    quantity: str = ""
    quantity_unit: str = ""
    expected_price_type: str = ""
    expected_tax_status: str = ""
    notes: str = ""


@dataclass
class ScrapeResult:
    product: Product
    checked_at: str
    result_state: str = "ok"
    field_status: dict[str, str] = field(default_factory=dict)
    price: float | None = None
    currency: str = "JPY"
    price_type: str = ""
    tax_status: str = ""
    unit_price: float | None = None
    unit_price_basis: str = ""
    stock_status: str = ""
    delivery_status: str = ""
    dispatch_status: str = ""
    raw_price_text: str = ""
    raw_stock_text: str = ""
    raw_delivery_text: str = ""
    title: str = ""
    error: str = ""
    source_checked_at: str = ""

    def to_latest_row(self) -> dict[str, object]:
        p = self.product
        return {
            "checked_at": self.checked_at,
            "result_state": self.result_state,
            "field_status": json.dumps(self.field_status, ensure_ascii=False, sort_keys=True),
            "group": p.group,
            "category": p.category,
            "site": p.site,
            "name": p.name,
            "product_code": p.product_code,
            "manufacturer": p.manufacturer,
            "manufacturer_code": p.manufacturer_code,
            "jan": p.jan,
            "cas": p.cas,
            "grade": p.grade,
            "capacity": p.capacity,
            "unit": p.unit,
            "quantity": p.quantity,
            "quantity_unit": p.quantity_unit,
            "price": self.price if self.price is not None else "",
            "currency": self.currency,
            "price_type": self.price_type,
            "tax_status": self.tax_status,
            "unit_price": self.unit_price if self.unit_price is not None else "",
            "unit_price_basis": self.unit_price_basis,
            "stock_status": self.stock_status,
            "delivery_status": self.delivery_status,
            "dispatch_status": self.dispatch_status,
            "raw_price_text": self.raw_price_text,
            "raw_stock_text": self.raw_stock_text,
            "raw_delivery_text": self.raw_delivery_text,
            "title": self.title,
            "url": p.url,
            "error": self.error,
            "source_checked_at": self.source_checked_at,
        }


def get_jst() -> dt.tzinfo:
    try:
        return ZoneInfo("Asia/Tokyo")
    except ZoneInfoNotFoundError:
        return dt.timezone(dt.timedelta(hours=9), name="JST")


def now_jst() -> dt.datetime:
    return dt.datetime.now(get_jst()).replace(microsecond=0)


def parse_bool(value: str | None, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on", "はい", "有効"}


def normalize_site(value: str) -> str:
    key = (value or "").strip().lower()
    return SITE_ALIASES.get(key, key)


def normalize_text(value: str) -> str:
    value = html.unescape(value or "")
    value = re.sub(r"<script\b.*?</script>", " ", value, flags=re.IGNORECASE | re.DOTALL)
    value = re.sub(r"<style\b.*?</style>", " ", value, flags=re.IGNORECASE | re.DOTALL)
    value = re.sub(r"<[^>]+>", " ", value)
    value = value.replace("\u3000", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def clean_fragment(value: str) -> str:
    value = normalize_text(value)
    value = re.sub(r"\s+", " ", value)
    return value.strip(" :-|/　")


def parse_number(value: str) -> float | None:
    if not value:
        return None
    normalized = value.translate(str.maketrans("０１２３４５６７８９，．", "0123456789,."))
    normalized = re.sub(r"[^\d.]", "", normalized.replace(",", ""))
    if not normalized:
        return None
    try:
        return float(normalized)
    except ValueError:
        return None


def extract_title(html_text: str) -> str:
    match = re.search(r"<title[^>]*>(.*?)</title>", html_text, flags=re.IGNORECASE | re.DOTALL)
    if match:
        return clean_fragment(match.group(1))
    og = re.search(
        r'<meta[^>]+property=["\']og:title["\'][^>]+content=["\']([^"\']+)["\']',
        html_text,
        flags=re.IGNORECASE,
    )
    return clean_fragment(og.group(1)) if og else ""


def extract_json_ld_prices(html_text: str) -> list[tuple[float, str, str]]:
    found: list[tuple[float, str, str]] = []
    blocks = re.findall(
        r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
        html_text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    for block in blocks:
        block = html.unescape(block.strip())
        try:
            payload = json.loads(block)
        except json.JSONDecodeError:
            continue
        for item in walk_json(payload):
            if not isinstance(item, dict):
                continue
            if "price" not in item:
                continue
            price = parse_number(str(item.get("price", "")))
            if price is None:
                continue
            currency = str(item.get("priceCurrency") or item.get("currency") or "JPY")
            found.append((price, currency, "json-ld"))
    return found


def walk_json(value: object) -> Iterable[object]:
    yield value
    if isinstance(value, dict):
        for child in value.values():
            yield from walk_json(child)
    elif isinstance(value, list):
        for child in value:
            yield from walk_json(child)


def first_regex(text: str, patterns: list[str]) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return clean_fragment(match.group(1))
    return ""


def price_from_context(text: str, labels: list[str]) -> tuple[float | None, str]:
    price_pattern = r"(?:(?:￥|¥)\s*([0-9０-９][0-9０-９,，\.．]*)|([0-9０-９][0-9０-９,，\.．]*)\s*円)"
    for label in labels:
        pattern = rf"{label}.{{0,80}}?{price_pattern}"
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            raw = match.group(0)
            price = parse_number(match.group(1) or match.group(2))
            if price is not None:
                return price, clean_fragment(raw)
    match = re.search(price_pattern, text)
    if match:
        return parse_number(match.group(1) or match.group(2)), clean_fragment(match.group(0))
    return None, ""


def infer_tax_status(text: str, fallback: str = "") -> str:
    if fallback:
        return fallback
    if re.search(r"税込|消費税込|tax included", text, flags=re.IGNORECASE):
        return "税込"
    if re.search(r"税抜|税別|本体価格|tax excluded", text, flags=re.IGNORECASE):
        return "税別"
    return "不明"


def calculate_unit_price(product: Product, price: float | None) -> tuple[float | None, str]:
    if price is None:
        return None, ""
    quantity = parse_number(product.quantity)
    unit = product.quantity_unit.strip()
    if quantity is None or quantity <= 0 or not unit:
        return None, ""
    return round(price / quantity, 4), f"1 {unit}"


def fetch_url(url: str, timeout: int, retries: int) -> str:
    last_error: Exception | None = None
    for attempt in range(retries + 1):
        request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                charset = response.headers.get_content_charset() or "utf-8"
                return response.read().decode(charset, errors="replace")
        except (urllib.error.URLError, TimeoutError, OSError) as exc:
            last_error = exc
            if attempt < retries:
                time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(str(last_error))


def parse_common(product: Product, html_text: str, checked_at: str) -> ScrapeResult:
    text = normalize_text(html_text)
    result = ScrapeResult(product=product, checked_at=checked_at, title=extract_title(html_text))

    json_prices = extract_json_ld_prices(html_text)
    if json_prices:
        result.price, result.currency, result.raw_price_text = json_prices[0]
        result.field_status["price"] = "ok"
    else:
        price, raw = price_from_context(text, [r"価格", r"販売価格", r"税込", r"税抜"])
        result.price = price
        result.raw_price_text = raw
        result.field_status["price"] = "ok" if price is not None else "not_found"

    result.tax_status = infer_tax_status(text, product.expected_tax_status)
    result.price_type = product.expected_price_type or "公開価格"
    result.stock_status = first_regex(
        text,
        [
            r"(在庫(?:状況)?[^。|,，]{0,40})",
            r"(販売終了[^。|,，]{0,30})",
            r"(欠品[^。|,，]{0,30})",
            r"(お取り寄せ[^。|,，]{0,30})",
        ],
    )
    result.delivery_status = first_regex(
        text,
        [
            r"((?:お届け|納期|到着予定)[^。|,，]{0,60})",
            r"((?:当日|翌日|明日)[^。|,，]{0,40}(?:お届け|出荷|配送))",
        ],
    )
    result.dispatch_status = first_regex(text, [r"((?:出荷|発送)[^。|,，]{0,60})"])
    result.raw_stock_text = result.stock_status
    result.raw_delivery_text = result.delivery_status
    return finalize_result(result)


def parse_monotaro(product: Product, html_text: str, checked_at: str) -> ScrapeResult:
    text = normalize_text(html_text)
    result = parse_common(product, html_text, checked_at)
    if result.price is None:
        price, raw = price_from_context(text, [r"販売価格", r"価格", r"注文コード"])
        result.price = price
        result.raw_price_text = raw
    result.price_type = product.expected_price_type or "販売価格"
    result.tax_status = infer_tax_status(text, product.expected_tax_status or "税別")
    return finalize_result(result)


def parse_askul(product: Product, html_text: str, checked_at: str) -> ScrapeResult:
    text = normalize_text(html_text)
    result = parse_common(product, html_text, checked_at)
    if result.price is None:
        price, raw = price_from_context(text, [r"税込", r"販売価格", r"価格"])
        result.price = price
        result.raw_price_text = raw
    result.price_type = product.expected_price_type or "販売価格"
    result.delivery_status = result.delivery_status or first_regex(
        text,
        [r"((?:お届け|お届け日|配送)[^。|,，]{0,60})"],
    )
    return finalize_result(result)


def parse_kanto(product: Product, html_text: str, checked_at: str) -> ScrapeResult:
    text = normalize_text(html_text)
    result = parse_common(product, html_text, checked_at)
    price, raw = price_from_context(text, [r"定価", r"価格"])
    if price is not None:
        result.price = price
        result.raw_price_text = raw
    result.price_type = product.expected_price_type or "定価"
    result.tax_status = infer_tax_status(text, product.expected_tax_status)
    result.stock_status = result.stock_status or first_regex(text, [r"(在庫[^。|,，]{0,60})"])
    return finalize_result(result)


def parse_wako(product: Product, html_text: str, checked_at: str) -> ScrapeResult:
    text = normalize_text(html_text)
    result = parse_common(product, html_text, checked_at)
    price, raw = price_from_context(text, [r"希望納入価格", r"キャンペーン価格", r"価格"])
    if price is not None:
        result.price = price
        result.raw_price_text = raw
    result.price_type = product.expected_price_type or (
        "希望納入価格" if "希望納入価格" in text else "公開価格"
    )
    result.tax_status = infer_tax_status(text, product.expected_tax_status)
    result.stock_status = result.stock_status or first_regex(
        text,
        [r"((?:東日本|西日本|在庫)[^。|,，]{0,80})"],
    )
    return finalize_result(result)


def finalize_result(result: ScrapeResult) -> ScrapeResult:
    if result.price is None:
        result.field_status["price"] = "not_found"
    else:
        result.field_status["price"] = "ok"
    result.field_status["stock"] = "ok" if result.stock_status else "not_found"
    result.field_status["delivery"] = "ok" if result.delivery_status else "not_found"
    result.unit_price, result.unit_price_basis = calculate_unit_price(result.product, result.price)
    missing_important = result.price is None
    result.result_state = "partial" if missing_important else "ok"
    return result


PARSERS = {
    "monotaro": parse_monotaro,
    "askul": parse_askul,
    "kanto": parse_kanto,
    "wako": parse_wako,
}


def product_from_row(row: dict[str, str]) -> Product:
    site = normalize_site(row.get("site", ""))
    name = row.get("name", "").strip()
    url = row.get("url", "").strip()
    category = row.get("category", "").strip()
    group = row.get("group", "").strip()
    if not group:
        group = row.get("compare_group", "").strip()
    if not group:
        group = name
    return Product(
        enabled=parse_bool(row.get("enabled"), default=True),
        group=group,
        category=category,
        site=site,
        name=name,
        url=url,
        product_code=row.get("product_code", "").strip(),
        manufacturer=row.get("manufacturer", "").strip(),
        manufacturer_code=row.get("manufacturer_code", "").strip(),
        jan=row.get("jan", "").strip(),
        cas=row.get("cas", "").strip(),
        grade=row.get("grade", "").strip(),
        capacity=row.get("capacity", "").strip(),
        unit=row.get("unit", "").strip(),
        quantity=row.get("quantity", "").strip(),
        quantity_unit=row.get("quantity_unit", "").strip(),
        expected_price_type=row.get("expected_price_type", "").strip(),
        expected_tax_status=row.get("expected_tax_status", "").strip(),
        notes=row.get("notes", "").strip(),
    )


def load_products(path: Path) -> list[Product]:
    if not path.exists():
        example_path = path.with_name("products.example.csv")
        guidance = [
            f"products file not found: {path}",
            "",
            "Create your local products.csv before running the monitor.",
        ]
        if example_path.exists():
            guidance.extend(
                [
                    f"Example template: {example_path}",
                    "",
                    "Windows command:",
                    f"  copy {example_path.name} {path.name}",
                ]
            )
        else:
            guidance.extend(
                [
                    "You can also create a starter file with:",
                    f"  python monitor.py --init-sample --products {path.name}",
                ]
            )
        raise FileNotFoundError("\n".join(guidance))
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        products = [product_from_row(row) for row in reader]
    return [p for p in products if p.enabled and p.url and p.site]


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, rows: list[dict[str, object]], columns: list[str], append: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    exists = path.exists() and path.stat().st_size > 0
    mode = "a" if append else "w"
    with path.open(mode, encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        if not append or not exists:
            writer.writeheader()
        for row in rows:
            writer.writerow(row)


def previous_success_by_key(history_rows: list[dict[str, str]]) -> dict[tuple[str, str, str], dict[str, str]]:
    previous: dict[tuple[str, str, str], dict[str, str]] = {}
    for row in history_rows:
        if row.get("result_state") not in {"ok", "partial"}:
            continue
        if not row.get("price"):
            continue
        key = (row.get("site", ""), row.get("product_code", ""), row.get("url", ""))
        previous[key] = row
    return previous


def add_history_deltas(
    latest_rows: list[dict[str, object]], history_rows: list[dict[str, str]]
) -> list[dict[str, object]]:
    previous = previous_success_by_key(history_rows)
    rows: list[dict[str, object]] = []
    for row in latest_rows:
        out = dict(row)
        key = (str(row.get("site", "")), str(row.get("product_code", "")), str(row.get("url", "")))
        prev = previous.get(key)
        price = parse_number(str(row.get("price", "")))
        prev_price = parse_number(prev.get("price", "")) if prev else None
        out["previous_success_price"] = prev_price if prev_price is not None else ""
        if price is not None and prev_price is not None:
            diff = price - prev_price
            out["price_diff"] = round(diff, 4)
            out["price_diff_percent"] = round((diff / prev_price) * 100, 4) if prev_price else ""
        else:
            out["price_diff"] = ""
            out["price_diff_percent"] = ""
        rows.append(out)
    return rows


def build_comparison_rows(latest_rows: list[dict[str, object]], checked_at: str) -> list[dict[str, object]]:
    groups: dict[str, list[dict[str, object]]] = {}
    for row in latest_rows:
        group = str(row.get("group", "")).strip()
        if not group:
            continue
        groups.setdefault(group, []).append(row)

    comparisons: list[dict[str, object]] = []
    for group, rows in groups.items():
        candidates = [r for r in rows if parse_number(str(r.get("price", ""))) is not None]
        if len(candidates) < 2:
            continue
        comparable = [
            r
            for r in candidates
            if str(r.get("currency", "JPY")) == "JPY"
            and str(r.get("tax_status", "")) in {"税別", "税込"}
            and str(r.get("result_state", "")) in {"ok", "partial"}
        ]
        if not comparable:
            comparisons.append(
                {
                    "checked_at": checked_at,
                    "group": group,
                    "category": rows[0].get("category", ""),
                    "comparison_note": "比較対象はありますが、税区分または価格がそろっていません",
                    "candidate_count": len(candidates),
                }
            )
            continue
        best = min(
            comparable,
            key=lambda r: parse_number(str(r.get("unit_price", "")))
            if parse_number(str(r.get("unit_price", ""))) is not None
            else parse_number(str(r.get("price", ""))) or 10**18,
        )
        comparisons.append(
            {
                "checked_at": checked_at,
                "group": group,
                "category": best.get("category", ""),
                "best_site": best.get("site", ""),
                "best_name": best.get("name", ""),
                "best_price": best.get("price", ""),
                "best_unit_price": best.get("unit_price", ""),
                "currency": best.get("currency", ""),
                "price_type": best.get("price_type", ""),
                "tax_status": best.get("tax_status", ""),
                "comparison_note": "同じgroup内の公開価格のみで比較。送料、契約価格、実購入可否は含みません",
                "candidate_count": len(candidates),
            }
        )
    return comparisons


def build_error_rows(latest_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    errors: list[dict[str, object]] = []
    for row in latest_rows:
        state = str(row.get("result_state", ""))
        field_status = str(row.get("field_status", ""))
        if state == "ok" and "not_found" not in field_status:
            continue
        failed_field = "page" if state == "error" else "field"
        reason = str(row.get("error", "")) or field_status
        errors.append(
            {
                "checked_at": row.get("checked_at", ""),
                "site": row.get("site", ""),
                "name": row.get("name", ""),
                "url": row.get("url", ""),
                "failed_field": failed_field,
                "reason": reason,
            }
        )
    return errors


def scrape_product(product: Product, checked_at: str, timeout: int, retries: int) -> ScrapeResult:
    parser = PARSERS.get(product.site, parse_common)
    try:
        html_text = fetch_url(product.url, timeout=timeout, retries=retries)
        return parser(product, html_text, checked_at)
    except Exception as exc:
        return ScrapeResult(
            product=product,
            checked_at=checked_at,
            result_state="error",
            field_status={"page": "error"},
            error=str(exc),
        )


def autosize_sheet(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")


def add_sheet(wb: Workbook, title: str, columns: list[str], rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])
    autosize_sheet(ws)


def write_workbook(
    path: Path,
    latest_rows: list[dict[str, object]],
    history_rows: list[dict[str, object]],
    compare_rows: list[dict[str, object]],
    error_rows: list[dict[str, object]],
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "最新一覧", LATEST_COLUMNS, latest_rows)
    add_sheet(wb, "販売店比較", COMPARE_COLUMNS, compare_rows)
    add_sheet(wb, "価格履歴", HISTORY_COLUMNS, history_rows)
    add_sheet(wb, "取得エラー", ERROR_COLUMNS, error_rows)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=str(path.parent)) as tmp:
        tmp_path = Path(tmp.name)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, path)
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_{now_jst().strftime('%Y%m%d_%H%M%S')}{path.suffix}")
        wb.save(fallback)
        return fallback
    finally:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


def ensure_sample_products(path: Path) -> None:
    if path.exists():
        return
    rows = [
        {
            "enabled": "true",
            "group": "kimwipes-s200",
            "category": "lab_consumable",
            "site": "monotaro",
            "name": "Kimwipes S-200 Monotaro sample",
            "url": "https://www.monotaro.com/p/0582/8462/",
            "product_code": "05828462",
            "manufacturer": "Nippon Paper Crecia",
            "unit": "1 box",
            "quantity": "200",
            "quantity_unit": "sheet",
            "expected_price_type": "販売価格",
            "expected_tax_status": "税別",
            "notes": "URL and SKU must be checked before operational use",
        },
        {
            "enabled": "false",
            "group": "kimwipes-s200",
            "category": "lab_consumable",
            "site": "askul",
            "name": "Kimwipes S-200 Askul sample",
            "url": "https://www.askul.co.jp/",
            "manufacturer": "Nippon Paper Crecia",
            "unit": "1 box",
            "quantity": "200",
            "quantity_unit": "sheet",
            "expected_price_type": "販売価格",
            "notes": "Replace with the exact product page and enable",
        },
        {
            "enabled": "false",
            "group": "reagent-example",
            "category": "reagent",
            "site": "kanto",
            "name": "Kanto reagent sample",
            "url": "https://cica-web.kanto.co.jp/",
            "expected_price_type": "定価",
            "notes": "Replace with exact Cica-Web product page and enable",
        },
        {
            "enabled": "false",
            "group": "reagent-example",
            "category": "reagent",
            "site": "wako",
            "name": "Wako reagent sample",
            "url": "https://labchem-wako.fujifilm.com/jp/",
            "expected_price_type": "希望納入価格",
            "notes": "Replace with exact Wako product page and enable",
        },
    ]
    write_csv(path, rows, PRODUCT_COLUMNS)


def run(args: argparse.Namespace) -> int:
    products_path = Path(args.products)
    if args.init_sample:
        ensure_sample_products(products_path)
        print(f"Sample products file is ready: {products_path}")
        return 0

    checked_at_dt = (
        dt.datetime.fromisoformat(args.checked_at)
        if args.checked_at
        else now_jst()
    )
    if checked_at_dt.tzinfo is None:
        checked_at_dt = checked_at_dt.replace(tzinfo=get_jst())
    checked_at = checked_at_dt.isoformat()

    products = load_products(products_path)
    output_dir = Path(args.output_dir)
    latest_path = output_dir / "latest.csv"
    history_path = output_dir / "history.csv"
    compare_path = output_dir / "comparison.csv"
    errors_path = output_dir / "errors.csv"
    workbook_path = output_dir / args.workbook

    lock_path = output_dir / ".monitor.lock"
    output_dir.mkdir(parents=True, exist_ok=True)
    try:
        lock_fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        os.write(lock_fd, str(os.getpid()).encode("ascii"))
        os.close(lock_fd)
    except FileExistsError:
        print(f"Another run appears to be active: {lock_path}", file=sys.stderr)
        return 2

    try:
        latest_results: list[ScrapeResult] = []
        for index, product in enumerate(products, start=1):
            if index > 1 and args.delay_seconds > 0:
                time.sleep(args.delay_seconds)
            print(f"[{index}/{len(products)}] {product.site}: {product.name}")
            latest_results.append(
                scrape_product(
                    product,
                    checked_at=checked_at,
                    timeout=args.timeout,
                    retries=args.retries,
                )
            )

        latest_rows = [result.to_latest_row() for result in latest_results]
        old_history = read_csv_rows(history_path)
        history_append_rows = add_history_deltas(latest_rows, old_history)
        all_history_rows = old_history + [
            {column: row.get(column, "") for column in HISTORY_COLUMNS} for row in history_append_rows
        ]
        compare_rows = build_comparison_rows(latest_rows, checked_at)
        error_rows = build_error_rows(latest_rows)

        write_csv(latest_path, latest_rows, LATEST_COLUMNS)
        write_csv(history_path, history_append_rows, HISTORY_COLUMNS, append=True)
        write_csv(compare_path, compare_rows, COMPARE_COLUMNS)
        write_csv(errors_path, error_rows, ERROR_COLUMNS)
        saved_workbook = write_workbook(
            workbook_path,
            latest_rows=latest_rows,
            history_rows=all_history_rows,
            compare_rows=compare_rows,
            error_rows=error_rows,
        )

        print(f"Latest CSV: {latest_path}")
        print(f"History CSV: {history_path}")
        print(f"Comparison CSV: {compare_path}")
        print(f"Errors CSV: {errors_path}")
        print(f"Excel workbook: {saved_workbook}")
        return 0 if not any(row.get("result_state") == "error" for row in latest_rows) else 1
    finally:
        try:
            lock_path.unlink()
        except OSError:
            pass


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Monitor public product price, stock, and delivery pages and export CSV/Excel files."
    )
    parser.add_argument("--products", default=DEFAULT_PRODUCTS, help="Product configuration CSV path.")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="Directory for CSV and Excel outputs.")
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, help="Excel workbook filename.")
    parser.add_argument("--timeout", type=int, default=20, help="HTTP timeout seconds per request.")
    parser.add_argument("--retries", type=int, default=1, help="Retry count after a failed request.")
    parser.add_argument("--delay-seconds", type=float, default=2.0, help="Delay between product requests.")
    parser.add_argument("--checked-at", default="", help="Override check timestamp in ISO format.")
    parser.add_argument("--init-sample", action="store_true", help="Create a sample products.csv if missing.")
    return parser


if __name__ == "__main__":
    sys.exit(run(build_parser().parse_args()))
